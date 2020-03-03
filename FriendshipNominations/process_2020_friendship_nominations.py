import openpyxl
import numpy as np
import pandas as pd
from person import Person

def extract_nominations(_row):

	# Initialize the person
	person_obj = Person(_row[0])

	print(person_obj.id)

	# Extract the nomination ids and top friends index
	start_idx = 1
	end_idx = start_idx + 10
	friend_noms = _row[start_idx:end_idx]
	top3_noms = _row[end_idx:(end_idx + 10)]

	person_obj.add_nom_list(friend_noms)
	person_obj.add_top3_list(top3_noms)

	# Create the nomination index (for top friend)
	top3_nom = []

	# Add checks
	# 1. Top index corresponds to actual friend and not null
	# 2. ClosestF index exists (e.g. not 3 and only 2 top3)

	# Iterate through each friendship nomination
	for i, nom_id in enumerate(friend_noms):

		if nom_id and nom_id > 1:
			person_obj.add_nomination(nom_id)

			# If equivalent top3_noms is 1, then close friend
			if top3_noms[i] == 1:
				person_obj.add_top3(nom_id)
				top3_nom.append(nom_id)

	# Add the top nomination
	closest_friend_idx = _row[end_idx + 10]
	if top3_nom and closest_friend_idx:

		# Get closest friend
		try:
			closest_friend = top3_nom[closest_friend_idx - 1]
			person_obj.add_closest(closest_friend)
		except IndexError:
			print("ERROR:\tClosest friend idx (%d) for %d not in top 3 nom!" % (closest_friend_idx, person_obj.id))

	return person_obj


def get_study_participants(_friendship_nominations_file):

	# Open the xls document
	wb = openpyxl.load_workbook(_friendship_nominations_file)

	# Assume only one sheet
	ws = wb.active

	# Get the header
	header = [x.value for x in ws[1]]

	# Initialize the person_dict
	person_dict = {}
	person_list = []

	# Iterate through each person
	for r in ws.iter_rows(min_row = 2):

		# Skip line if blank
		if not r[0].value:
			continue

		# Get the row into a list
		row_list = [v.value for v in r]

		person_obj = extract_nominations(row_list)
		person_list.append(person_obj.id)

		person_dict[person_obj.id] = person_obj

	return person_list, person_dict, header

def check_reciprocal_nom(_person_dict):

	# Iterate through each person
	for id, pobj in _person_dict.items():

		pobj.check_nom_list_reciprocated(_person_dict)
		pobj.check_top3_list_reciprocated(_person_dict)
		pobj.check_closest_reciprocated(_person_dict)

if __name__ == "__main__":

	data_file = "../data/NCSSM/sample Friend Noms 2.27.20.xlsx"

	student_list, students, header = get_study_participants(data_file)

	check_reciprocal_nom(students)

	# Add recip to each column
	header_output = [header[0]]
	for h in header[1:]:
		header_output.append(h)
		header_output.append(h + "_recip")

	with open(file = "output.txt", mode = "w") as fO:

		fO.write("\t".join(header_output) + "\n")
		for s in student_list:
			p_obj = students[s]
			out_str = "\t".join([str(s), p_obj.write_nom_row(), p_obj.write_top3_row(), p_obj.write_closest()])
			fO.write(out_str + "\n")

'''

def GetPersonDict(arg_f_in):
    
    # Open the xls document
    wb = openpyxl.load_workbook(arg_f_in)
    ws = wb["Sheet1"]


    # Initialize the personDict


def InitOutputDF(arg_person_dict):
    
    # Set the column names
    outHeader = ["ID", "ID_bafmean"]
    
    # Iterate through each nomination
    for i in range(1, 11):
        # Set the nom
        nom_str = "Nom_" + str(i)
        outHeader += [nom_str, nom_str + "_bafmean", nom_str + "_recip"]
    
    # Add in the top3 nom
    for i in range(1, 4):
        # Set the nom
        nom_str = "Top3_Nom_" + str(i)
        outHeader += [nom_str, nom_str + "_bafmean", nom_str + "_recip"]
    
    # Add in the top
    nom_str = "Top_Nom"
    outHeader += [nom_str, nom_str + "_bafmean", nom_str + "_recip"]
    
    # Initialize the dataframe
    out_df = pd.DataFrame(index = range(0, len(arg_person_dict)),
                          columns = outHeader
                         )
    
    # Fill in the ID and bafmean
    r_idx = 0
    for stID, p_obj in arg_person_dict.items():
        
        # Get the bafMean
        bafMean = p_obj.get_belief()
        
        # Insert into the dataframe
        out_df.loc[r_idx, ["ID", "ID_bafmean"]] = [stID, bafMean]
        r_idx += 1
    
    # Sort on id
    out_df.sort_values(by = "ID", axis = 0, ascending = True, inplace = True)
    
    return out_df

def UpdateNom(arg_df, arg_person_dict):
    
    for i in arg_df.index:
        
        # Obtain the ID
        stID = arg_df.loc[i, "ID"]
        
        # Get the nomination set
        nom_set = arg_person_dict[stID].get_nominations()
        
        # Initialize the nomination index
        nom_idx = 1
        for nom in nom_set:
            
            # Get the nom column prefix
            nomColPre = "Nom_" + str(nom_idx)
            
            # Enter nom into data frame
            arg_df.loc[i, nomColPre] = nom
            
            # Check if nom is in the person_dict
            if nom not in arg_person_dict:
                nom_bafmean = np.nan
                nom_recip =  9
            else:
                nom_bafmean = arg_person_dict[nom].get_belief()
                nom_recip = 1 if arg_person_dict[nom].is_id_in_nom(stID) else 0
                
            # Update the nom bafmean and nom recip
            arg_df.loc[i, nomColPre + "_bafmean"] = nom_bafmean
            arg_df.loc[i, nomColPre + "_recip"] = nom_recip
    
            # Increment the nom_idx
            nom_idx += 1
    
    return arg_df

def UpdateTop3(arg_df, arg_person_dict):
    
    for i in arg_df.index:
        
        # Obtain the ID
        stID = arg_df.loc[i, "ID"]
        
        # Get the top3 set
        nom_set = arg_person_dict[stID].get_top3()
        
        if nom_set == None:
            continue
        
        # Initialize the nomination index
        nom_idx = 1
        for nom in nom_set:
            
            # Get the nom column prefix
            nomColPre = "Top3_Nom_" + str(nom_idx)
            
            # Enter nom into data frame
            arg_df.loc[i, nomColPre] = nom
            
            # Check if nom is in the person_dict
            if nom not in arg_person_dict:
                nom_bafmean = np.nan
                nom_recip =  9
            else:
                nom_bafmean = arg_person_dict[nom].get_belief()
                
                # Assign 0, 1, 2 if not reciprocated, reciprocated but not in top3, or in top3
                if arg_person_dict[nom].is_id_in_top3(stID):
                    nom_recip = 2
                elif arg_person_dict[nom].is_id_in_nom(stID):
                    nom_recip = 1
                else:
                    nom_recip = 0
                
            # Update the nom bafmean and nom recip
            arg_df.loc[i, nomColPre + "_bafmean"] = nom_bafmean
            arg_df.loc[i, nomColPre + "_recip"] = nom_recip
    
            # Increment the nom_idx
            nom_idx += 1
    
    return arg_df

def UpdateTop(arg_df, arg_person_dict):
    
    for i in arg_df.index:
        
        # Obtain the ID
        stID = arg_df.loc[i, "ID"]
        
        # Get the top 
        topNom = arg_person_dict[stID].get_top()
        
        if topNom == None:
            continue
        
        # Get the nom column prefix
        nomColPre = "Top_Nom"
        # Enter nom into data frame
        arg_df.loc[i, nomColPre] = topNom
            
        # Check if nom is in the person_dict
        if topNom not in arg_person_dict:
            nom_bafmean = np.nan
            nom_recip =  9
        else:
            nom_bafmean = arg_person_dict[topNom].get_belief()
            
            # Assign 0, 1, 2 if not reciprocated, reciprocated but not in top3, or in top3
            if arg_person_dict[topNom].is_id_in_top(stID):
                nom_recip = 3
            elif arg_person_dict[topNom].is_id_in_top3(stID):
                nom_recip = 2
            elif arg_person_dict[topNom].is_id_in_nom(stID):
                nom_recip = 1
            else:
                nom_recip = 0
            
        # Update the nom bafmean and nom recip
        arg_df.loc[i, nomColPre + "_bafmean"] = nom_bafmean
        arg_df.loc[i, nomColPre + "_recip"] = nom_recip
    
    return arg_df
    
    
# Set the input file
#f_in = "c:/usr/cygwin64/home/jab112/P_data3/fake reciprocity file 9.3.18.xlsx"
#f_out = "c:/usr/cygwin64/home/jab112/P_data3/reciprocity_output.txt"
f_in = "Z:/NCSSM Study 2018/Data/Friendship Nominat1Gions 8.8.18.xlsx"
f_out = "Z:/NCSSM Study 2018/Data/Friendship Reciprocity Output.txt"


# Obtain each student data from the reciprocity matrix
person_dict = GetPersonDict(f_in)

# Setup the output dataframe
nom_df = InitOutputDF(person_dict)

# Update the nominations for each person
nom_df = UpdateNom(nom_df, person_dict)

# Update the top3 nominations for each person
nom_df = UpdateTop3(nom_df, person_dict)

# Update the top nomination for each person
nom_df = UpdateTop(nom_df, person_dict)

# Write to tab-delimited file
nom_df.to_csv(f_out, sep = "\t", na_rep = "", header = True, index = False)
'''
