# -*- coding: utf-8 -*-
"""
Created on Wed Jun 27 21:16:15 2018

@author: jab112
"""

import openpyxl
import pandas as pd
import numpy as np
import glob

fO = open("Class_All.txt", mode = "w")
fO.write("%s\t%s\t%s\t%s\t%s\t%s\n" % ("StudentID", "Gender", "Given_Mean", "Given_Std", "Received_Mean", "Received_Std"))

f_xls = "c:/usr/cygwin64/home/jab112/P_data2/Socalla_Item_1_20160711.xlsx"
wb = openpyxl.load_workbook(f_xls)
for sheet_name in wb.get_sheet_names()[0:1]:
    ws = wb[sheet_name]
    
    # Initialize the iStart and iEnd
    colA = [c.value for c in ws["A"]]
    iStart = colA.index("ID") + 1
    iEnd = ws.max_row
    
    # Find the first blank space after the numbers
    for i,v in enumerate(colA[iStart:]):
        if v == None:
            iEnd = iStart + i
            break
    
    
    
    colA = colA[iStart:iEnd]
    gender_idx = pd.Series([c.value for c in ws["B"][iStart:iEnd]], index = colA)
    
    # Construct the output dataframe
    output_df = pd.DataFrame(np.column_stack((colA, gender_idx)), 
                              columns = ["StudentID", "Gender"], 
                              index = np.arange(len(colA))
                             )
    
    # Add in the additional columns
    addl_columns = ["Given_mean", "Given_ClassGender_mean", "Given_ClassGender_std", "Given_ClassGender_zscore", "Received_mean", "Received_ClassGender_mean", "Received_ClassGender_std", "Received_ClassGender_zscore"]
    output_df = output_df.reindex(columns = output_df.columns.tolist() + addl_columns, fill_value = 0)
    
    score_matrix = []
    
    for r in ws.iter_rows(min_row = iStart + 1, max_row = iEnd, min_col = iStart + 1, max_col = iEnd):
        score_matrix.append([c.value for c in r])
    
    df = pd.DataFrame(score_matrix, index = colA, columns = colA)
    
    
    # Iterate through each id
    for studentID in colA:
        #print(studentID)
    
        if studentID in [2309, 2707, 2708, 4114, 4124, 4621, 4819]:
            continue
    
        for gr in ["Given", "Received"]:    
    
            if gr == "Given":
                score = df.loc[studentID,:]
                colname = "Given_mean"
            elif gr == "Received":
                score = df.loc[:,studentID]
                colname = "Received_mean"
                
            # Exclude the missing values
            score = score[score != 9]
        
            if len(score) > 0:
                meanScore = score.mean()
            else:
                meanScore = 9
    
            # Select the output column
            output_df.loc[output_df["StudentID"] == studentID, colname] = meanScore
    
    # Get the class-gender mean and std
    for g in [0, 1]:
    
        for gr in ["Given", "Received"]:
                       
            # Set the idx
            idx = (output_df["Gender"] == g) & (output_df[gr + "_mean"] != 9)
        
            output_df.loc[idx, gr + "_ClassGender_mean"] = output_df.loc[idx, gr + "_mean"].mean()
            output_df.loc[idx, gr + "_ClassGender_std"] = output_df.loc[idx, gr + "_mean"].std()

    # Set the missings to 9!        
    # Get rid of the chain indexing from below


    for stIdx in output_df.index:

        for gr in ["Given", "Received"]:
            
            # Calculate z-score
            curMean = output_df.iloc[stIdx][gr + "_mean"]
            classMean = output_df.iloc[stIdx][gr + "_ClassGender_mean"]
            stDev = output_df.iloc[stIdx][gr + "_ClassGender_std"]
            
            z_score = (curMean - classMean) / stDev
            
            output_df.iloc[stIdx][gr + "_ClassGender_zscore"] = z_score
        
fO.close()
'''

  
        if len(given_score.unique()) > 1:        
            given_std = given_score.std()
        else:
            given_std = 9
        
    
        #fO.write("%d\t%d\t%.4f\t%.4f\t%.4f\t%.4f\n" % (studentID, gender_idx[studentID], given_mean, given_std, received_mean, received_std))
    

d = "c:/usr/cygwin64/home/jab112/P_data/excel_files/"
d_out = "c:/usr/cygwin64/home/jab112/P_data/overall/"

for item in range(1,13):

    f_peer = "Peer provisions item %s" % item
    
    f_xls = glob.glob(d + "Peer provisions item " + str(item) + "_*.xlsx")
    print(f_xls)
    
    
    #print("Processing %s..." % f)
    
    wb = openpyxl.load_workbook(f_xls[0])
    
    fO = open(d_out + "output_item " + str(item) + ".txt", mode = "w")
    fO.write("\t".join(["StudentID", "Given", "Received"]) + "\n")
    
    for wksheetname in wb.get_sheet_names():
    
        
        print(wksheetname)
        
        if not wksheetname.startswith("Class "):
            continue
        
        ws = wb[wksheetname]
    
        # Get the header
        r = [c.value for c in ws["1"]]
        iStart = r.index("ID") + 1
        iEnd = r.index("Gender Proportion")
        student_id = r[iStart:iEnd]
        
        # Get the gender
        gender = [c.value for c in ws["2"][iStart:iEnd]]
        
        # Make this a dict
        studentGender = {}
        
        # Initialize given and received dict
        stuGiv = {}
        stuRec = {}
        
        for i,st in enumerate(student_id):
            studentGender[st] = gender[i]
        
        for r in ws.iter_rows(min_row = 3, max_row = iEnd, min_col = 1, max_col = iEnd):
            
            # Get the id
            stID = r[0].value
            
            row = [c.value for c in r[2:]]
            
            # Initialize total and counter
            tot = 0
            ct = 0
            
            # Get the student gender
            stGender = studentGender[stID]
            
            for i,v in enumerate(row):
                #if v != 9 and gender[i] != stGender:
                if v != 9:    
                    tot += v
                    ct += 1
            
            # Get the final number
            if ct == 0:
                final_val = 9
            else:
                final_val = tot / ct
            
            stuGiv[stID] = final_val    
            
            
                
        for c in ws.iter_cols(min_row = 1, max_row = iEnd, min_col = 3, max_col = iEnd):
            
            # Get the id
            stID = c[0].value
            
            col = [r.value for r in c[2:]]
            
            # Initialize total and counter
            tot = 0
            ct = 0
            
            # Get the student gender
            stGender = studentGender[stID]
            
            for i,v in enumerate(col):
                #if v != 9 and gender[i] != stGender:
                if v != 9:    
                    tot += v
                    ct += 1
            
            # Get the final number
            if ct == 0:
                final_val = 9
            else:
                final_val = tot / ct
            
            stuRec[stID] = final_val    
        
            
        for st in student_id:
            fO.write(str(st) + "\t" + str(stuGiv[st]) + "\t" + str(stuRec[st]) + "\n")

    fO.close()
'''
