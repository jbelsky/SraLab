# -*- coding: utf-8 -*-
"""
Created on Wed Jun 27 21:16:15 2018

@author: jab112
"""

import openpyxl
import pandas as pd
import numpy as np
import argparse
import itertools
from collections import OrderedDict

def GetStuAvgScore(arg_out_df, arg_score_df, arg_id, arg_gender, arg_type):

	# Get the given or received score
	score = pd.Series(dtype = np.int64)
	if arg_type == "Given":
		score = arg_score_df.loc[arg_id]
	elif arg_type == "Received":
		score = arg_score_df[arg_id]

	# Get scores != 9
	# See if there needs to be a gender filter
	iScore = score != 9
	if arg_gender == "ByGender":
		iScore = iScore & (arg_out_df["Gender"] == arg_out_df.loc[arg_id, "Gender"])
	elif arg_gender == "CrossGender":
		iScore = iScore & (arg_out_df["Gender"] != arg_out_df.loc[arg_id, "Gender"])

	# Subset on the score
	score = score[iScore]
	scoreMean = score.mean() if len(score) > 0 else 9

	arg_out_df.loc[arg_id, "mean"] = scoreMean

def GetClassAvgScore(arg_out_df, arg_gender):

	# Obtain the indices where there is a mean
	iScore = arg_out_df["mean"] != 9

	# Set no mean score std and mean to 9
	if any(~iScore):
		arg_out_df.loc[~iScore, "ClassGender_mean"] = 9
		arg_out_df.loc[~iScore, "ClassGender_std"] = 9

	# Iterate through each gender case
	if arg_gender == "AllGender":
		arg_out_df["ClassGender_mean"] = arg_out_df["mean"].mean()
		arg_out_df["ClassGender_std"] = classStd = arg_out_df["mean"].std()
	elif arg_gender == "ByGender":
		# Iterate through each gender separately
		for g in arg_out_df["Gender"].unique():

			# Set the idx
			iScoreGender = iScore & (arg_out_df["Gender"] == g)

			arg_out_df.loc[iScoreGender, "ClassGender_mean"] = arg_out_df.loc[iScoreGender,"mean"].mean()
			arg_out_df.loc[iScoreGender, "ClassGender_std"] = arg_out_df.loc[iScoreGender,"mean"].std()
	elif arg_gender == "CrossGender":
		# Iterate through each gender separately
		for g in arg_out_df["Gender"].unique():

			# Set the idx
			iScoreGender = iScore & (arg_out_df["Gender"] != g)

			arg_out_df.loc[iScoreGender, "ClassGender_mean"] = arg_out_df.loc[iScoreGender,"mean"].mean()
			arg_out_df.loc[iScoreGender, "ClassGender_std"] = arg_out_df.loc[iScoreGender,"mean"].std()





parser = argparse.ArgumentParser()
parser.add_argument("input_excel", help = "the input Socalla Excel sheet")
# parser.add_argument("output_text", help = "the output text file")
args = parser.parse_args()
wb = openpyxl.load_workbook(args.input_excel)

#for i in range(15, 53):
for i in range(15, 16):

	sheet_name = "Class " + str(i)
	print(sheet_name)
	ws = wb[sheet_name]

	# Find the indices
	columnA = np.array([c.value for c in ws["A"]])
	iStuIDs = np.where(~((columnA == None) | (columnA == "ID")))[0]
	stuIDs = pd.Index(columnA[iStuIDs])

	# Get the gender (Column B)
	iStart = iStuIDs[0]
	iEnd = iStuIDs[-1]
	genders = pd.Series([c.value for c in ws["B"][iStart:(iEnd + 1)]], index = stuIDs, dtype = np.int16)

	# Create the pandas 2-D matrix
	score_matrix = []
	for r in ws.iter_rows(min_row = iStart + 1, max_row = iEnd + 1, min_col = iStart + 1, max_col = iEnd + 1):
		score_matrix.append([c.value for c in r])
	score_df = pd.DataFrame(score_matrix, index = stuIDs, columns = stuIDs, dtype = np.int16)

	# Construct the output dataframe
	outTemplate_df = pd.DataFrame(np.column_stack((stuIDs, genders)),
								      columns = ["StudentID", "Gender"],
									   index = stuIDs
								      )

	# Add in the additional columns
	addl_columns = ["mean", "ClassGender_mean", "ClassGender_std", "ClassGender_zscore"]
	outTemplate_df = outTemplate_df.reindex(columns = outTemplate_df.columns.tolist() + addl_columns, fill_value = 0)

	# Need to patch colA, gender_idx, output_df, and df if any of the
	STUDENT_IDS_SKIP = [2309, 2707, 2708, 4114, 4124, 4621, 4819]
	if any(stuIDs.isin(STUDENT_IDS_SKIP)):
		stuID_drop = stuIDs[stuIDs.isin(STUDENT_IDS_SKIP)]
		outTemplate_df.drop(stuID_drop, inplace = True)
		score_df.drop(index = stuID_drop, columns = stuID_drop, inplace = True)

	outTemplate_df["Gender"].iloc[0:2] = 7

	# Ensure that there aren't any unknown genders
	if(any(~outTemplate_df["Gender"].isin([0,1]))):
		iUnknownGender = ~outTemplate_df["Gender"].isin([0,1])
		for s in outTemplate_df.index[np.where(iUnknownGender)[0]]:
			print("ERROR: %d gender is %d (must be '0' or '1')!" % (s, outTemplate_df.loc[s,"Gender"]))
		break

	# Create the output for each type [AllGender, ByGender, CrossGender]
	genderComps = ["AllGender", "ByGender", "CrossGender"]
	analysisTypes = ["Given", "Received"]
	data_d = OrderedDict()

	for gComp, aType in itertools.product(genderComps, analysisTypes):

		output_df = outTemplate_df.copy()

		# Get the average score for each student
		for stuID in output_df.index:

			# Get the student average score
			GetStuAvgScore(output_df, score_df, stuID, gComp, aType)

		# Get the class gender mean and std

		data_d[(gComp, aType)] = output_df

'''
		# Get the class-gender mean and std
		output_df.loc[, ("ClassGender_mean", "ClassGender_std")]

	# Get the class-gender mean and std
	for g in [0, 1]:

		for gr in ["Given", "Received"]:

			# Set the idx
			idx = (output_df["Gender"] == g) & (output_df[gr + "_mean"] != 9)

			output_df.loc[idx, gr + "_ClassGender_mean"] = output_df.loc[idx, gr + "_mean"].mean()
			output_df.loc[idx, gr + "_ClassGender_std"] = output_df.loc[idx, gr + "_mean"].std()

			# Set the missing means and std to 9
			idx = (output_df["Gender"] == g) & (output_df[gr + "_mean"] == 9)
			output_df.loc[idx, gr + "_ClassGender_mean"] = 9
			output_df.loc[idx, gr + "_ClassGender_std"] = 9


	# Set the missings to 9!
	# Get rid of the chain indexing from below


	for stIdx in output_df.index:

		for gr in ["Given", "Received"]:

			# Calculate z-score
			curMean = output_df.loc[stIdx, gr + "_mean"]
			classMean = output_df.loc[stIdx, gr + "_ClassGender_mean"]
			stDev = output_df.loc[stIdx, gr + "_ClassGender_std"]

			#if sheet_name == "Class 23":
			#	print(str(output_df.loc[stIdx, "StudentID"]) + "\t" + "\t".join([str(curMean),str(classMean),str(stDev)]))

			# Get the z_score if not missing
			if curMean == 9:
				z_score = 9
			else:
				z_score = (curMean - classMean) / stDev

			output_df.loc[stIdx,gr + "_ClassGender_zscore"] = z_score

	if sheet_name == "Class 15":
		output_df.to_csv(args.output_text, mode = "w", sep = "\t", header = True, index = False)
	else:
		output_df.to_csv(args.output_text, mode = "a", sep = "\t", header = False, index = False)

'''