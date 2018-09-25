# -*- coding: utf-8 -*-
"""
Created on Wed Jun 27 21:16:15 2018

@author: jab112
"""

import openpyxl
import pandas as pd
import numpy as np
import argparse
    
parser = argparse.ArgumentParser()
parser.add_argument("input_excel", help = "the input Socalla Excel sheet")
parser.add_argument("output_text", help = "the output text file")
args = parser.parse_args()
wb = openpyxl.load_workbook(args.input_excel)

#for i in range(15, 53):
for i in range(23, 24):
	sheet_name = "Class " + str(i)
	print(sheet_name)
	ws = wb[sheet_name]
	
	# Initialize the iStart and iEnd
	colA = [c.value for c in ws["A"]]
	iStart = colA.index("ID") + 1
	iEnd = ws.max_row
	
	# Find the first blank space after the numbers
	for i,v in enumerate(colA[iStart:]):
		if v == None:
			iEnd = iStart + i
			break
	
	
	
	colA = colA[iStart:iEnd]
	gender_idx = pd.Series([c.value for c in ws["B"][iStart:iEnd]], index = colA)
	
	# Construct the output dataframe
	output_df = pd.DataFrame(np.column_stack((colA, gender_idx)), 
							  columns = ["StudentID", "Gender"], 
							  index = np.arange(len(colA))
							 )
	
	# Add in the additional columns
	addl_columns = ["Given_mean", "Given_ClassGender_mean", "Given_ClassGender_std", "Given_ClassGender_zscore", "Received_mean", "Received_ClassGender_mean", "Received_ClassGender_std", "Received_ClassGender_zscore"]
	output_df = output_df.reindex(columns = output_df.columns.tolist() + addl_columns, fill_value = 0)
	
	score_matrix = []
	
	for r in ws.iter_rows(min_row = iStart + 1, max_row = iEnd, min_col = iStart + 1, max_col = iEnd):
		score_matrix.append([c.value for c in r])
	
	df = pd.DataFrame(score_matrix, index = colA, columns = colA)
	
	# Need to patch colA, gender_idx, output_df, and df if any of the 
	if studentID in [2309, 2707, 2708, 4114, 4124, 4621, 4819]:
		output_df.loc[output_df["StudentID"] == studentID, output_df.columns.tolist()[2:]] = 9
		continue
	
	
	
	# Iterate through each id
	for studentID in colA:
		#print(studentID)


		for gr in ["Given", "Received"]:
	
			if gr == "Given":
				score = df.loc[studentID,:]
				colname = "Given_mean"
			elif gr == "Received":
				score = df.loc[:,studentID]
				colname = "Received_mean"
			
			# Obtain the current student gender
			gender = gender_idx[studentID]
			
			# Create the score subset
			#scoreSubset_idx = score != 9
			#scoreSubset_idx = (score != 9) & (gender_idx == gender)
			scoreSubset_idx = (score != 9) & (gender_idx != gender)

			# Subset on the score
			score = score[scoreSubset_idx]
		
			if len(score) > 0:
				meanScore = score.mean()
			else:
				meanScore = 9
	
			# Select the output column
			output_df.loc[output_df["StudentID"] == studentID, colname] = meanScore
	
	# Get the class-gender mean and std
	for g in [0, 1]:
	
		for gr in ["Given", "Received"]:
					   
			# Set the idx
			idx = (output_df["Gender"] == g) & (output_df[gr + "_mean"] != 9)
		
			output_df.loc[idx, gr + "_ClassGender_mean"] = output_df.loc[idx, gr + "_mean"].mean()
			output_df.loc[idx, gr + "_ClassGender_std"] = output_df.loc[idx, gr + "_mean"].std()

			# Set the missing means and std to 9
			idx = (output_df["Gender"] == g) & (output_df[gr + "_mean"] == 9)
			output_df.loc[idx, gr + "_ClassGender_mean"] = 9
			output_df.loc[idx, gr + "_ClassGender_std"] = 9
   

	# Set the missings to 9!		
	# Get rid of the chain indexing from below


	for stIdx in output_df.index:

		for gr in ["Given", "Received"]:
			
			# Calculate z-score
			curMean = output_df.loc[stIdx, gr + "_mean"]
			classMean = output_df.loc[stIdx, gr + "_ClassGender_mean"]
			stDev = output_df.loc[stIdx, gr + "_ClassGender_std"]
			
			#if sheet_name == "Class 23":
			#	print(str(output_df.loc[stIdx, "StudentID"]) + "\t" + "\t".join([str(curMean),str(classMean),str(stDev)]))
			
			# Get the z_score if not missing
			if curMean == 9:
				z_score = 9
			else:
				z_score = (curMean - classMean) / stDev
			
			output_df.loc[stIdx,gr + "_ClassGender_zscore"] = z_score
	
	if sheet_name == "Class 15":
		output_df.to_csv(args.output_text, mode = "w", sep = "\t", header = True, index = False)
	else:
		output_df.to_csv(args.output_text, mode = "a", sep = "\t", header = False, index = False)
