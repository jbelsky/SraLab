import openpyxl
import pandas as pd
import os
import glob
import re
import sys
from collections import OrderedDict

import classmatrix
import OParkStudent

def GetProvisions(ARG_PROV_DIR):

	# Initialize the peerProvision matrix
	provDict = {}

	# Obtain the number of peer provision matrices
	provXlsxFiles = glob.glob(ARG_PROV_DIR + "/*.xlsx")
	provXlsxFiles.sort()

	# Get the peer provision item -> file dict
	p1 = re.compile("Peer provisions (item \d{1,2})_.*\.xlsx")
	p2 = re.compile("(Item \d{1,2})-.*\.xlsx")

	for provXlsxFile in provXlsxFiles:

		print(provXlsxFile)

		m1 = p1.match(os.path.basename(provXlsxFile))
		m2 = p2.match(os.path.basename(provXlsxFile))
		if m1:
			item = m1.group(1)
		elif m2:
			item = m2.group(1)
		else:
			print("Could not extract 'item' from provision Excel filename!")
			sys.exit(1)

		# Initialize the dict for the item
		provDict[item] = {}

		# Open the Excel file
		wb = openpyxl.load_workbook(provXlsxFile)
		sheetnames = wb.sheetnames
		sheetnames.sort()

		for sn in sheetnames:

			if "Class" not in sn:
				continue

			data_df, gender_s = classmatrix.GetDataMatrix(wb[sn])

			# Enter into dict
			provDict[item][sn] = data_df

	return provDict

friendship_file = "C:/Programs/cygwin64/home/jab112/github/SraLab/data/Orland_Park/Friendship_Nominations/Socallb item 5 Unlimited Friend Noms 5.17.17.xlsx"

# Open the Excel file
wb = openpyxl.load_workbook(friendship_file)

friendship_df, gender_s = classmatrix.GetDataMatrix(wb["Class 15"])

students = OrderedDict()

for i in gender_s.items():

        stID = i[0]
        gender = i[1]

        students[stID] = OParkStudent.OParkStudent(stID, gender)
