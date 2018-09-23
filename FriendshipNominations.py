# -*- coding: utf-8 -*-
"""
Created on Sun Aug 26 14:38:05 2018

@author: jab112
"""

import openpyxl
import numpy as np
import pandas as pd
import string

class Person:
    
    def __init__(self, arg_id, arg_belief):
        self.id = arg_id
        self.belief = arg_belief
        self.nominations = set()
        self.top3 = None
        self.top = None
    
    def add_nomination(self, arg_id):
        self.nominations.add(arg_id)
    
    def check_if_nominated(self, arg_id):
        if arg_id in self.nominations:
            return True
        else:
            return False
    
    def set_top_three(self, arg_id_arr):
        if len(arg_id_arr) == 0:
            self.top3 = None
        else:
            self.top3 = set(arg_id_arr)
    
    def set_top(self, arg_id):
        self.top = arg_id

    def get_belief(self):
        return self.belief
    
    def get_nominations(self):
        return self.nominations
    
    def get_top3(self):
        return self.top3
    
    def get_top(self):
        return self.top
    
    def is_id_in_nom(self, arg_id):
        return True if arg_id in self.nominations else False
        
    def is_id_in_top3(self, arg_id):
        if self.top3 == None:
            return False
        elif arg_id in self.top3:
            return True
        else:
            return False
    
    def is_id_in_top(self, arg_id):
        if self.top == None:
            return False
        elif arg_id == self.top:
            return True
        else:
            return False

def GetPersonDict(arg_f_in):
    
    # Open the xls document
    wb = openpyxl.load_workbook(arg_f_in)
    ws = wb["Sheet1"]

    # Get the header
    headerRow = [h.value for h in ws["1"]]

    # Initialize the personDict
    person_dict = {}

    # Iterate through each person
    for r in ws.iter_rows(min_row = 2):
    
        if r[0].value == None:
            continue
    
        # Get the row into a list
        rowList = [v.value for v in r]
        
        # Check if the bafmean is defined
        bafMean = rowList[headerRow.index("bafmean")]
        if bafMean == None:
            bafMean = np.nan
    
        # Initialize the person
        person_obj = Person(rowList[headerRow.index("StudentID")], bafMean)
    
        # Add in all the nominations
        for i in range(headerRow.index("Nom_1"), headerRow.index("Nom_10") + 1):
            if rowList[i] != None:
                person_obj.add_nomination(rowList[i])
    
        # Get the top 3 nominations
        top3_list = rowList[headerRow.index("Step1"):(headerRow.index("Step3") + 1)]
        
        # Get the top nomination index
        topNom_idx = rowList[headerRow.index("Nom_One")]
    
        # Check if topNom_idx is valid
        if topNom_idx in range(1, 4):
            topNom = top3_list[topNom_idx - 1]
        else:
            topNom = None
        
        # Enter the top3_list and the topNom into the Person
        top3_arr = np.array(top3_list)
        top3_arr = top3_arr[top3_arr != None]
        person_obj.set_top_three(top3_arr)
        person_obj.set_top(topNom)
            
        # Enter into dict
        person_dict[person_obj.id] = person_obj
    
    return person_dict        

def InitOutputDF(arg_person_dict):
    
    # Set the column names
    outHeader = ["ID", "ID_bafmean"]
    
    # Iterate through each nomination
    for i in range(1, 11):
        # Set the nom
        nom_str = "Nom_" + str(i)
        outHeader += [nom_str, nom_str + "_bafmean", nom_str + "_recip"]
    
    # Add in the top3 nom
    for i in range(1, 4):
        # Set the nom
        nom_str = "Top3_Nom_" + str(i)
        outHeader += [nom_str, nom_str + "_bafmean", nom_str + "_recip"]
    
    # Add in the top
    nom_str = "Top_Nom"
    outHeader += [nom_str, nom_str + "_bafmean", nom_str + "_recip"]
    
    # Initialize the dataframe
    out_df = pd.DataFrame(index = range(0, len(arg_person_dict)),
                          columns = outHeader
                         )
    
    # Fill in the ID and bafmean
    r_idx = 0
    for stID, p_obj in arg_person_dict.items():
        
        # Get the bafMean
        bafMean = p_obj.get_belief()
        
        # Insert into the dataframe
        out_df.loc[r_idx, ["ID", "ID_bafmean"]] = [stID, bafMean]
        r_idx += 1
    
    # Sort on id
    out_df.sort_values(by = "ID", axis = 0, ascending = True, inplace = True)
    
    return out_df

def UpdateNom(arg_df, arg_person_dict):
    
    for i in arg_df.index:
        
        # Obtain the ID
        stID = arg_df.loc[i, "ID"]
        
        # Get the nomination set
        nom_set = arg_person_dict[stID].get_nominations()
        
        # Initialize the nomination index
        nom_idx = 1
        for nom in nom_set:
            
            # Get the nom column prefix
            nomColPre = "Nom_" + str(nom_idx)
            
            # Enter nom into data frame
            arg_df.loc[i, nomColPre] = nom
            
            # Check if nom is in the person_dict
            if nom not in arg_person_dict:
                nom_bafmean = np.nan
                nom_recip =  9
            else:
                nom_bafmean = arg_person_dict[nom].get_belief()
                nom_recip = 1 if arg_person_dict[nom].is_id_in_nom(stID) else 0
                
            # Update the nom bafmean and nom recip
            arg_df.loc[i, nomColPre + "_bafmean"] = nom_bafmean
            arg_df.loc[i, nomColPre + "_recip"] = nom_recip
    
            # Increment the nom_idx
            nom_idx += 1
    
    return arg_df

def UpdateTop3(arg_df, arg_person_dict):
    
    for i in arg_df.index:
        
        # Obtain the ID
        stID = arg_df.loc[i, "ID"]
        
        # Get the top3 set
        nom_set = arg_person_dict[stID].get_top3()
        
        if nom_set == None:
            continue
        
        # Initialize the nomination index
        nom_idx = 1
        for nom in nom_set:
            
            # Get the nom column prefix
            nomColPre = "Top3_Nom_" + str(nom_idx)
            
            # Enter nom into data frame
            arg_df.loc[i, nomColPre] = nom
            
            # Check if nom is in the person_dict
            if nom not in arg_person_dict:
                nom_bafmean = np.nan
                nom_recip =  9
            else:
                nom_bafmean = arg_person_dict[nom].get_belief()
                
                # Assign 0, 1, 2 if not reciprocated, reciprocated but not in top3, or in top3
                if arg_person_dict[nom].is_id_in_top3(stID):
                    nom_recip = 2
                elif arg_person_dict[nom].is_id_in_nom(stID):
                    nom_recip = 1
                else:
                    nom_recip = 0
                
            # Update the nom bafmean and nom recip
            arg_df.loc[i, nomColPre + "_bafmean"] = nom_bafmean
            arg_df.loc[i, nomColPre + "_recip"] = nom_recip
    
            # Increment the nom_idx
            nom_idx += 1
    
    return arg_df

def UpdateTop(arg_df, arg_person_dict):
    
    for i in arg_df.index:
        
        # Obtain the ID
        stID = arg_df.loc[i, "ID"]
        
        # Get the top 
        topNom = arg_person_dict[stID].get_top()
        
        if topNom == None:
            continue
        
        # Get the nom column prefix
        nomColPre = "Top_Nom"
        # Enter nom into data frame
        arg_df.loc[i, nomColPre] = topNom
            
        # Check if nom is in the person_dict
        if topNom not in arg_person_dict:
            nom_bafmean = np.nan
            nom_recip =  9
        else:
            nom_bafmean = arg_person_dict[topNom].get_belief()
            
            # Assign 0, 1, 2 if not reciprocated, reciprocated but not in top3, or in top3
            if arg_person_dict[topNom].is_id_in_top(stID):
                nom_recip = 3
            elif arg_person_dict[topNom].is_id_in_top3(stID):
                nom_recip = 2
            elif arg_person_dict[topNom].is_id_in_nom(stID):
                nom_recip = 1
            else:
                nom_recip = 0
            
        # Update the nom bafmean and nom recip
        arg_df.loc[i, nomColPre + "_bafmean"] = nom_bafmean
        arg_df.loc[i, nomColPre + "_recip"] = nom_recip
    
    return arg_df
    
    
# Set the input file
#f_in = "c:/usr/cygwin64/home/jab112/P_data3/fake reciprocity file 9.3.18.xlsx"
#f_out = "c:/usr/cygwin64/home/jab112/P_data3/reciprocity_output.txt"
f_in = "Z:/NCSSM Study 2018/Data/Friendship Nominations 8.8.18.xlsx"
f_out = "Z:/NCSSM Study 2018/Data/Friendship Reciprocity Output.txt"


# Obtain each student data from the reciprocity matrix
person_dict = GetPersonDict(f_in)

# Setup the output dataframe
nom_df = InitOutputDF(person_dict)

# Update the nominations for each person
nom_df = UpdateNom(nom_df, person_dict)

# Update the top3 nominations for each person
nom_df = UpdateTop3(nom_df, person_dict)

# Update the top nomination for each person
nom_df = UpdateTop(nom_df, person_dict)

# Write to tab-delimited file
nom_df.to_csv(f_out, sep = "\t", na_rep = "", header = True, index = False)
