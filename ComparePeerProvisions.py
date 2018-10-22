# -*- coding: utf-8 -*-
"""
Created on Wed Jun 27 21:16:15 2018

Updated: 2018-09-30

@author: jab112
"""

import openpyxl
import pandas as pd
import numpy as np
import argparse
import os
import glob
import re

def GetDataMatrix(_ws):

	# Find the relevant indices
	columnA = np.array([c.value for c in _ws["A"]])

	# Find the last numeric index after the None, 'ID', data types
	if any(columnA == "ID"):
		iStart = np.where(columnA == "ID")[0][0] + 1
	else:
		# Start on first integer
		for i,v in enumerate(columnA):
			if type(v) == int:
				iStart = i
				break

	for i,v in enumerate(columnA[iStart:]):

		if type(v) != int:
			iEnd = i - 1
			break

	# Get the row list of values (will serve as input to DataFrame)
	rowList = []
	for row in _ws.iter_rows(min_row = 1, max_row = iEnd, min_col = 1, max_col = iEnd):
		rowList.append([c.value for c in row])

	# Return a data frame
	return pd.DataFrame(rowList)

parser = argparse.ArgumentParser()
parser.add_argument("excel_dir1", help = "Excel directory 1")
parser.add_argument("excel_dir2", help = "Excel directory 1")
parser.add_argument("-o", "--output", action = "store", default = "compare_excel_dirs.txt", type = str)
args = parser.parse_args()

# Open the output version
fO = open(args.output, mode = "w")
fO.write("FileName\tSheetName\tComparisonStatus\n")

# Obtain the Excel files in each directory
dir1_files = glob.glob(args.excel_dir1 + "/*.xlsx")
dir2_files = glob.glob(args.excel_dir2 + "/*.xlsx")

dir1_files.sort()
dir2_files.sort()

for i,xls1 in enumerate(dir1_files[0:1]):

	xls2 = dir2_files[i]

	# Ensure that the item numbers match between the two versions
	p = re.compile(".*Peer provisions (item \d{1,2})_.*\.xlsx")
	xls1_item = p.search(xls1).group(1)
	xls2_item = p.search(xls2).group(1)

	if xls1_item != xls2_item:
		print("WARNING: The comparison files have different item numbers, skipping...\n\t%s\n\t%s\n" % (xls1, xls2))
		continue
	else:
		print("%s\t%s" % (xls1_item, xls2_item))

	wb1 = openpyxl.load_workbook(xls1)
	wb2 = openpyxl.load_workbook(xls2)

	print("Processing %s..." % os.path.basename(xls1))

	for sheet_name in wb1.get_sheet_names():

		if "Class" not in sheet_name:
			print("Skipping %s..." % sheet_name)
			continue

		print(sheet_name)

		df1 = GetDataMatrix(wb1[sheet_name])
		df2 = GetDataMatrix(wb2[sheet_name])

		if df1.equals(df2):
			fO.write("%s\t%s\t%s\n" % (os.path.basename(xls1), sheet_name, "Equivalent"))
		else:
			fO.write("%s\t%s\t%s\n" % (os.path.basename(xls1), sheet_name, "DIFFERENT"))
			print("%s\t%s\t%s\n" % (os.path.basename(xls1), sheet_name, "DIFFERENT"))

fO.close()