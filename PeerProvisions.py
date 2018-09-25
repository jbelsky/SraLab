# -*- coding: utf-8 -*-
"""
Created on Wed Jun 27 21:16:15 2018

@author: jab112
"""

import openpyxl
import numpy as np
import glob

d = "c:/usr/cygwin64/home/jab112/P_data/excel_files/"
d_out = "c:/usr/cygwin64/home/jab112/P_data/overall/"

for item in range(1,13):

    f_peer = "Peer provisions item %s" % item
    
    f_xls = glob.glob(d + "Peer provisions item " + str(item) + "_*.xlsx")
    print(f_xls)
    
    
    #print("Processing %s..." % f)
    
    wb = openpyxl.load_workbook(f_xls[0])
    
    fO = open(d_out + "output_item " + str(item) + ".txt", mode = "w")
    fO.write("\t".join(["StudentID", "Given", "Received"]) + "\n")
    
    for wksheetname in wb.get_sheet_names():
    
        
        print(wksheetname)
        
        if not wksheetname.startswith("Class "):
            continue
        
        ws = wb[wksheetname]
    
        # Get the header
        r = [c.value for c in ws["1"]]
        iStart = r.index("ID") + 1
        iEnd = r.index("Gender Proportion")
        student_id = r[iStart:iEnd]
        
        # Get the gender
        gender = [c.value for c in ws["2"][iStart:iEnd]]
        
        # Make this a dict
        studentGender = {}
        
        # Initialize given and received dict
        stuGiv = {}
        stuRec = {}
        
        for i,st in enumerate(student_id):
            studentGender[st] = gender[i]
        
        for r in ws.iter_rows(min_row = 3, max_row = iEnd, min_col = 1, max_col = iEnd):
            
            # Get the id
            stID = r[0].value
            
            row = [c.value for c in r[2:]]
            
            # Initialize total and counter
            tot = 0
            ct = 0
            
            # Get the student gender
            stGender = studentGender[stID]
            
            for i,v in enumerate(row):
                #if v != 9 and gender[i] != stGender:
                if v != 9:    
                    tot += v
                    ct += 1
            
            # Get the final number
            if ct == 0:
                final_val = 9
            else:
                final_val = tot / ct
            
            stuGiv[stID] = final_val    
            
            
                
        for c in ws.iter_cols(min_row = 1, max_row = iEnd, min_col = 3, max_col = iEnd):
            
            # Get the id
            stID = c[0].value
            
            col = [r.value for r in c[2:]]
            
            # Initialize total and counter
            tot = 0
            ct = 0
            
            # Get the student gender
            stGender = studentGender[stID]
            
            for i,v in enumerate(col):
                #if v != 9 and gender[i] != stGender:
                if v != 9:    
                    tot += v
                    ct += 1
            
            # Get the final number
            if ct == 0:
                final_val = 9
            else:
                final_val = tot / ct
            
            stuRec[stID] = final_val    
        
            
        for st in student_id:
            fO.write(str(st) + "\t" + str(stuGiv[st]) + "\t" + str(stuRec[st]) + "\n")

    fO.close()
