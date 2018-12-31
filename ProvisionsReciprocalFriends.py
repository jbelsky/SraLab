# -*- coding: utf-8 -*-
"""
Created on Thu Nov 15 17:07:00 2018

@author: BelskyJ
"""

import openpyxl
import pandas as pd
import os
import glob
import argparse
import re
from collections import OrderedDict

from orlandpark import classmatrix
from orlandpark import OParkClass
from orlandpark import OParkStudent

def GetPeerProvisions(_dir):

	# Initialize the peerProvision matrix
	classPeerProvisionByItem_dict = {}

	# Obtain the number of peer provision matrices
	peerProv_xlsx_files = glob.glob(_dir + "/*xlsx")
	peerProv_xlsx_files.sort()

	# Get the peer provision item -> file dict
	p = re.compile("Peer provisions item (\d{1,2})_.*\.xlsx")

	for f in peerProv_xlsx_files:

		print(f)

		item = int(p.match(os.path.basename(f)).group(1))
		classPeerProvisionByItem_dict[item] = {}

		pp_wb = openpyxl.load_workbook(f)

		for OPclass in pp_wb.sheetnames:

			if "Class" not in OPclass:
				continue

			data_df, gender_s = classmatrix.GetDataMatrix(pp_wb[OPclass])

			# Enter into dict
			classPeerProvisionByItem_dict[item][OPclass] = data_df

		pp_wb.close()

	return classPeerProvisionByItem_dict


def GetClassFriendshipForEachStudent(_friendship_df):

	# Initialize the class_friendship dict
	classFriendship_dict = OrderedDict()

	# Iterate through each row
	for i in _friendship_df.index:

		# Initialize the student series
		studentFriendship_srs = pd.Series(index = _friendship_df.columns, dtype = str)

		for j in _friendship_df.columns:

			given = _friendship_df.loc[i, j]
			received = _friendship_df.loc[j, i]

			if (given == 9) | (received == 9):
				studentFriendship_srs.loc[j] = "NA"
				continue

			# Get the given and received status
			isGiven = True if given == 1 else False
			isReceived = True if received == 1 else False

			if isGiven and isReceived:
				studentFriendship_srs.loc[j] = "reciprocated"
			elif isGiven:
				studentFriendship_srs.loc[j] = "given"
			elif isReceived:
				studentFriendship_srs.loc[j] = "received"
			else:
				studentFriendship_srs.loc[j] = "none"

		# Add to the dict
		classFriendship_dict[i] = studentFriendship_srs

	# Return the friendship matrix
	return classFriendship_dict

if __name__ == "__main__":

	parser = argparse.ArgumentParser()
	parser.add_argument("friendship_nom_file", help = "Friendship Nominations File (.xlsx)")
	parser.add_argument("peer_provisions_dir", help = "Directory containing peer provisions")
	parser.add_argument("-o", "--output", action = "store", default = "FriendshipPeerProvisionsByItemAnalysis.xlsx", type = str)
	args = parser.parse_args()

	# Get the peer provisions for each of the classes
	classPeerProvisionsByItem_dict = GetPeerProvisions(args.peer_provisions_dir)

	# Load in the friendship matrix
	wb = openpyxl.load_workbook(args.friendship_nom_file)
	class_sn = wb.sheetnames
	class_sn.sort()

	# Set up the friendship_matrix dict
	OPclass_dict = OrderedDict()

	for c in class_sn:

		print("Working on %s..." % c)

		# Import the nominations matrix
		friendship_df, gender_srs = classmatrix.GetDataMatrix(wb[c])

		# Get the friendship dict
		friendship_dict = GetClassFriendshipForEachStudent(friendship_df)

		# Set the OParkStudent list
		OPstudent_list = [OParkStudent.OParkStudent(i, gender_srs.loc[i]) for i in gender_srs.index]

		# Initialize an OPClass
		orlandParkClass = OParkClass.OParkClass(c, OPstudent_list, friendship_dict, classPeerProvisionsByItem_dict)

		# Iterate through each OParkStudent
		for ops in OPstudent_list:

			ops.set_oParkClass(orlandParkClass)
			ops.set_received_peerprov_df()
			ops.set_received_friendships()
			ops.set_provision_received_by_friendship_status()
			metric_df = ops.set_provision_stats()

		# Get the class item summary
		orlandParkClass.set_item_summary()

		# Store the item statistics in the dict
		OPclass_dict[c] = orlandParkClass.itemStat_odict

	wb.close()

	# Create the output workbook
	f_out = "FriendshipPeerProvisionsByItemAnalysis.xlsx"
	writer= pd.ExcelWriter(f_out)

	# Iterate through each item
	items = list(classPeerProvisionsByItem_dict.keys())
	items.sort()
	for i in items:

		# Get all the dataframes for the item in one list
		itemByClass = []
		for c, opc_d in OPclass_dict.items():
			itemByClass.append(opc_d[i])

		# Write the excel worksheet
		itemByClass_df = pd.concat(itemByClass, axis = 0, join = "inner")
		itemByClass_df.to_excel(writer, sheet_name = "Item_" + str(i), float_format = "%.4f", freeze_panes = (1, 1))

	writer.save()
