import openpyxl
from person import Person


def check_recip_nom(_cur_stu, _stu_dict):

    cur_stu = _cur_stu # type: Person
    stu_dict = _stu_dict # type: dict

    # Initialize the storage list
    out_l = []
    for friend_id in cur_stu.get_nominations().keys():
        out_l.append(friend_id)
        if friend_id not in stu_dict:
            recip_val = 9
        else:
            recip_val = cur_stu.check_reciprocal_friend(stu_dict[friend_id])
        out_l.append(recip_val)

    # Add "" values to ensure len(out_l) == 20
    if len(out_l) < 20:
        out_l += ["" for x in range(20 - len(out_l))]

    return out_l

def check_recip_top3(_cur_stu, _stu_dict):

    cur_stu = _cur_stu # type: Person
    stu_dict = _stu_dict # type: dict

    # Initialize the storage list
    out_l = []
    for friend_id in cur_stu.get_top3().keys():
        out_l.append(friend_id)
        if friend_id not in stu_dict:
            recip_val = 9
        else:
            recip_val = cur_stu.check_reciprocal_friend(stu_dict[friend_id])
        out_l.append(recip_val)

    # Add "" values to ensure len(out_l) == 20
    if len(out_l) < 6:
        out_l += ["" for x in range(6 - len(out_l))]

    return out_l

def check_recip_closest(_cur_stu, _stu_dict):

    cur_stu = _cur_stu # type: Person
    stu_dict = _stu_dict # type: dict

    closest = cur_stu.get_closest()
    if not closest:
        return ["", ""]
    elif closest not in stu_dict:
        return [closest, 9]
    else:
        recip_val = cur_stu.check_reciprocal_closest(stu_dict[closest])
        return [closest, recip_val]

data_file = "sample Friend Noms 2.27.20.xlsx"

# Open the xls document
wb = openpyxl.load_workbook(data_file)

# Assume only one sheet
ws = wb.active

# Get the header
header = [x.value for x in ws[1]]

# Initialize the person dict
p_dict = {}

# Iterate through each person
for r in ws.iter_rows(min_row=2):

    # Skip line if blank
    if not r[0].value:
        continue

    # Get the row into a list
    row_list = [v.value for v in r]

    p = Person(row_list[0])

    p.add_nominations(row_list)

    p_dict[p.get_id()] = p

# Iterate through each person to check reciprocals
with open("output.txt", mode = "w") as fO:

    # Write the header
    header = ["StudentID"]
    for i in range(1, 11):
        header += ["Friend" + str(i), "Friend" + str(i) + "_recip"]
    for i in range(1, 4):
        header += ["Top3_" + str(i), "Top3_" + str(i) + "_recip"]
    header += ["Closest", "Closest_recip"]
    fO.write("\t".join(header) + "\n")

    for student_id, student_obj in p_dict.items():

        out_list = [student_id]
        out_list += check_recip_nom(student_obj, p_dict)
        out_list += check_recip_top3(student_obj, p_dict)
        out_list += check_recip_closest(student_obj, p_dict)

        fO.write("\t".join([str(x) for x in out_list]) + "\n")
