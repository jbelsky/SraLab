import openpyxl

# Set the filename
input_file_name = "C:/Programs/cygwin64/home/jab112/github/SraLab/data/MultipleFriends/Friend_Nominations_1.11.20.xlsx"

# Open the Excel file
wb = openpyxl.load_workbook(input_file_name)
ws = wb.active

# Initialize the dict of studyIDs
ids = {}

for r in ws["A"][1:]:
	id = r.value
	if id not in ids:
		ids[id] = 0
	else:
		print("WARNING: duplicate id found:\t%s", id)

# Iterate through columns B:K
# If id matches an id in ids, increment by 1
for row in ws.iter_rows(min_row = 2, min_col = 2):
	for c in row:
		id = c.value
		if id in ids:
			ids[id] += 1

# Sort the keys
ids_keys = list(ids.keys())
ids_keys.sort()

# Print the output
with open("C:/Programs/cygwin64/home/jab112/github/SraLab/data/MultipleFriends/Friend_Nominations_1.11.20_summary.txt", mode = "w") as fO:
	fO.write("StudyID\tTotalNominations\n")
	for k in ids_keys:
		fO.write("%d\t%d\n" % (k, ids[k]))
